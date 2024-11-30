import face_recognition
import cv2
import numpy as np
from openpyxl import Workbook, load_workbook # type: ignore
from datetime import datetime, timedelta
import os
from tkinter import Tk, simpledialog, messagebox

# Initialisation de Tkinter
root = Tk()
root.withdraw()  # Masquer la fenêtre principale Tkinter

# Chemin du fichier Excel
excel_file_path = "Présence.xlsx"

# Charger les visages connus depuis le dossier images/
known_face_encodings = []
known_face_names = []

def load_known_faces(folder_path="images/"):
    for filename in os.listdir(folder_path):
        if filename.endswith((".jpg", ".jpeg", ".png")):
            image_path = os.path.join(folder_path, filename)
            image = face_recognition.load_image_file(image_path)
            face_encodings = face_recognition.face_encodings(image)
            if face_encodings:
                known_face_encodings.append(face_encodings[0])
                name = os.path.splitext(filename)[0]  # Utiliser le nom du fichier (sans extension) comme nom de la personne
                known_face_names.append(name)
                print(f"Visage chargé : {name}")
            else:
                print(f"Aucun visage trouvé dans {filename}")

# Charger toutes les images existantes dans le dossier "images/"
load_known_faces()

# Fonction pour marquer la présence dans le fichier Excel
def mark_attendance_in_excel(name, excel_file_path):
    try:
        # Charger le fichier Excel ou créer un nouveau fichier s'il n'existe pas
        if not os.path.exists(excel_file_path):
            wb = Workbook()
            ws = wb.active
            ws.title = "Attendance"
            ws.append(["Nom"] + [f"Pointage {i}" for i in range(1, 101)])  # Colonnes initiales
            wb.save(excel_file_path)

        wb = load_workbook(excel_file_path)
        ws = wb.active

        # Chercher la ligne correspondant au nom
        row = None
        for i, cell in enumerate(ws["A"], start=1):
            if cell.value == name:
                row = i
                break

        # Si le nom n'existe pas, ajouter une nouvelle ligne
        if row is None:
            row = ws.max_row + 1
            ws.cell(row=row, column=1, value=name)

        # Ajouter le pointage dans la première colonne libre
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        for col in range(2, ws.max_column + 2):
            if ws.cell(row=row, column=col).value is None:
                ws.cell(row=row, column=col, value=current_time)
                break

        wb.save(excel_file_path)
        print(f"Pointage enregistré pour {name} à {current_time}")
    except Exception as e:
        print(f"Erreur lors de l'écriture dans le fichier Excel : {e}")

# Fonction pour enregistrer une nouvelle personne
def register_new_face(face_encoding, frame):
    while True:
        name = simpledialog.askstring("Nom", "Personne inconnue détectée.\nVeuillez entrer le nom de la personne :")
        if not name:  # Si l'utilisateur annule ou laisse le champ vide
            print("Enregistrement annulé.")
            return None
        if name in known_face_names:
            # Si le nom existe déjà, afficher un message d'erreur
            messagebox.showerror("Nom déjà utilisé", f"Le nom '{name}' existe déjà.\nVeuillez entrer un autre nom.")
        else:
            # Ajouter le visage et enregistrer l'image
            known_face_encodings.append(face_encoding)
            known_face_names.append(name)
            image_path = f"images/{name}.jpg"
            cv2.imwrite(image_path, frame)
            print(f"Nouvelle personne enregistrée : {name}")
            return name

# Initialisation de la webcam
video_capture = cv2.VideoCapture(0)

# Initialisation de certaines variables
face_locations = []
face_encodings = []
face_names = []
process_this_frame = True

# Dictionnaire pour suivre les derniers pointages (éviter doublons avant 4 heures)
attendance_dict = {}

while True:
    ret, frame = video_capture.read()
    small_frame = cv2.resize(frame, (0, 0), fx=0.25, fy=0.25)
    rgb_small_frame = cv2.cvtColor(small_frame, cv2.COLOR_BGR2RGB)

    if process_this_frame:
        face_locations = face_recognition.face_locations(rgb_small_frame)
        face_encodings = face_recognition.face_encodings(rgb_small_frame, face_locations)

        face_names = []
        for face_encoding in face_encodings:
            matches = face_recognition.compare_faces(known_face_encodings, face_encoding)
            name = "Unknown"

            if known_face_encodings:
                face_distances = face_recognition.face_distance(known_face_encodings, face_encoding)
                best_match_index = np.argmin(face_distances) if face_distances.size > 0 else None

                if best_match_index is not None and matches[best_match_index]:
                    name = known_face_names[best_match_index]
            else:
                face_distances = []
                print("Aucun visage connu disponible pour comparaison.")

            if name == "Unknown":
                # Enregistrer une nouvelle personne si elle est inconnue
                new_name = register_new_face(face_encoding, frame)
                if new_name:
                    name = new_name

            face_names.append(name)

            if name != "Unknown":
                now = datetime.now()
                if name not in attendance_dict or now - attendance_dict[name] > timedelta(hours=4):
                    attendance_dict[name] = now
                    mark_attendance_in_excel(name, excel_file_path)

    process_this_frame = not process_this_frame

    # Afficher les résultats
    for (top, right, bottom, left), name in zip(face_locations, face_names):
        top *= 4
        right *= 4
        bottom *= 4
        left *= 4

        cv2.rectangle(frame, (left, top), (right, bottom), (0, 0, 255), 2)
        cv2.rectangle(frame, (left, bottom - 35), (right, bottom), (0, 0, 255), cv2.FILLED)
        cv2.putText(frame, name, (left + 6, bottom - 6), cv2.FONT_HERSHEY_DUPLEX, 1.0, (255, 255, 255), 1)

    cv2.imshow('Face Attendance', frame)

    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

video_capture.release()
cv2.destroyAllWindows()
